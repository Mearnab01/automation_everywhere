import pandas as pd
import os
from openpyxl import Workbook
from openpyxl import load_workbook

# Function to merge CSV files
def merge_csv_files(csv_files, output_file):
    combined_data = pd.concat([pd.read_csv(file) for file in csv_files])
    combined_data.to_csv(output_file, index=False)

# Function to merge Excel (XLSX) files
def merge_excel_files(excel_files, output_file):
    wb = Workbook()
    dest = wb.active

    for file in excel_files:
        src = load_workbook(file, read_only=True)
        for sheet in src:
            for row in sheet.iter_rows(values_only=True):
                dest.append(row)
    
    wb.save(output_file)

# Get the path from the user
data_path = input("Enter the path where the Excel (XLSX) or CSV files are located: ")

# Initialize lists to store file paths
csv_files = []
excel_files = []

# Iterate through files in the specified directory
for file in os.listdir(data_path):
    file_path = os.path.join(data_path, file)
    if file.endswith(".csv"):
        csv_files.append(file_path)
    elif file.endswith(".xlsx"):
        excel_files.append(file_path)

# Specify the output file path
output_file = os.path.join(data_path, "merged_data.xlsx")

# Merge CSV files if any
if csv_files:
    merge_csv_files(csv_files, output_file)

# Merge Excel (XLSX) files if any
if excel_files:
    merge_excel_files(excel_files, output_file)

print(f"Files in {data_path} merged into {output_file}")
